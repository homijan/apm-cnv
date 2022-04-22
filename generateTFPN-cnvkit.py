#!/usr/bin/env python3

import openpyxl
from cnvkitweightedcopynumber import cnvkitWeightedCN

ref_table_name = 'tables/CNV_final_R_totals.xlsx'
ref_book = openpyxl.load_workbook(ref_table_name, read_only=True)
ref_sheet = ref_book.active

cnvkit_cn_table_name = 'tables/CNV_copynumber_cnvkit.xlsx'
cnvkit_cn_book = openpyxl.load_workbook(cnvkit_cn_table_name)
cnvkit_cn_sheet = cnvkit_cn_book.active

cnvkit_TFPN_table_name = 'tables/CNV_TFPN_cnvkit.xlsx'
cnvkit_TFPN_book = openpyxl.load_workbook(cnvkit_TFPN_table_name)
cnvkit_TFPN_sheet = cnvkit_TFPN_book.active

Nrows = ref_sheet.max_row
Ncolumns = ref_sheet.max_column

colNames = {}
for i in range(Ncolumns):
  col = i + 1
  colName = ref_sheet.cell(1, col).value
  colNames[colName] = col

print(f'HSTAMP_Label FISH')
for i in range(118):#range(Nrows-1):
  row = i + 2 
  HLabel = ref_sheet.cell(row, colNames['HSTAMP_Label']).value
  callFile = f'cnvkit/results-cnn-tumor/Sample_{HLabel}-T1_Tumor.samtools.call.cns'
  w_cns = cnvkitWeightedCN(callFile)
  for w_cn in w_cns: 
    cn = w_cns[w_cn]
    kcna = f'k{w_cn}'
    fcna = f'f{w_cn}'
    tcna = f't{w_cn}'
    cnvkit_cn_sheet.cell(row, colNames[kcna]).value = cn
    cnvkit_cn_sheet.cell(row, colNames[fcna]).value = cn
    cnvkit_cn_sheet.cell(row, colNames[tcna]).value = cn
    # Clinical result of karyotype or fish
    TP = 'TP'; FP = 'FP'; FN = 'FN'; TN = 'TN'; NA = 'NA'
    #TP = 87; FP = 37; FN = 36; TN = 86; NA = 62
    tcna_value = ref_sheet.cell(row, colNames[tcna]).value
    # tcna of loss is t5q, t7q, etc, and gain is ttrisomy, hence tcna[1] = 't'
    if (tcna[1]=='t'):
      # Gain true-false-positive-negative logic
      if (tcna_value==1 and cn > 2.0):
        cnvkit_TFPN_sheet.cell(row, colNames[tcna]).value = TP 
      elif (tcna_value==0 and cn > 2.0):
        cnvkit_TFPN_sheet.cell(row, colNames[tcna]).value = FP
      elif (tcna_value==1 and cn <= 2.0):
        cnvkit_TFPN_sheet.cell(row, colNames[tcna]).value = FN
      elif (tcna_value==0 and cn <= 2.0):
        cnvkit_TFPN_sheet.cell(row, colNames[tcna]).value = TN
      else:
        cnvkit_TFPN_sheet.cell(row, colNames[tcna]).value = NA
    else:
      # Deletion true-false-positive-negative logic
      if (tcna_value==1 and cn < 2.0):
        cnvkit_TFPN_sheet.cell(row, colNames[tcna]).value = TP 
      elif (tcna_value==0 and cn < 2.0):
        cnvkit_TFPN_sheet.cell(row, colNames[tcna]).value = FP
      elif (tcna_value==1 and cn >= 2.0):
        cnvkit_TFPN_sheet.cell(row, colNames[tcna]).value = FN
      elif (tcna_value==0 and cn >= 2.0):
        cnvkit_TFPN_sheet.cell(row, colNames[tcna]).value = TN
      else:
        cnvkit_TFPN_sheet.cell(row, colNames[tcna]).value = NA

  print(f'{HLabel} {w_cns}')
# Save the updated excel tables
cnvkit_cn_book.save(cnvkit_cn_table_name)
cnvkit_TFPN_book.save(cnvkit_TFPN_table_name)
