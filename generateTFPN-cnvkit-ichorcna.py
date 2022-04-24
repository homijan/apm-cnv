#!/usr/bin/env python3

import openpyxl
import sys
from CNAdefs import *
from weightedcopynumber import weightedCN

# Clinical result of karyotype or fish
TP = 'TP'; FP = 'FP'; FN = 'FN'; TN = 'TN'; NA = 'NA'
def updateSheets(row, colNames, w_cns, cn_sheet, TFPN_sheet):
  for w_cn in w_cns: 
    cn = w_cns[w_cn]
    kcna = f'k{w_cn}'
    fcna = f'f{w_cn}'
    tcna = f't{w_cn}'
    cn_sheet.cell(row, colNames[kcna]).value = cn
    cn_sheet.cell(row, colNames[fcna]).value = cn
    cn_sheet.cell(row, colNames[tcna]).value = cn
    tcna_value = ref_sheet.cell(row, colNames[tcna]).value
    # tcna of loss is t5q, t7q, etc, and gain is t1q, ttrisomy8 , and ttrisomy12 (skip leading 't')
    if (tcna[1:]=='1q' or (tcna[1:]=='trisomy8' or tcna[1:]=='trisomy12')):
      # Gain true-false-positive-negative logic
      if (tcna_value==1 and cn > 2.0):
        TFPN_sheet.cell(row, colNames[tcna]).value = TP 
      elif (tcna_value==0 and cn > 2.0):
        TFPN_sheet.cell(row, colNames[tcna]).value = FP
      elif (tcna_value==1 and cn <= 2.0):
        TFPN_sheet.cell(row, colNames[tcna]).value = FN
      elif (tcna_value==0 and cn <= 2.0):
        TFPN_sheet.cell(row, colNames[tcna]).value = TN
      else:
        TFPN_sheet.cell(row, colNames[tcna]).value = NA
    else:
      # Deletion true-false-positive-negative logic
      if (tcna_value==1 and cn < 2.0):
        TFPN_sheet.cell(row, colNames[tcna]).value = TP 
      elif (tcna_value==0 and cn < 2.0):
        TFPN_sheet.cell(row, colNames[tcna]).value = FP
      elif (tcna_value==1 and cn >= 2.0):
        TFPN_sheet.cell(row, colNames[tcna]).value = FN
      elif (tcna_value==0 and cn >= 2.0):
        TFPN_sheet.cell(row, colNames[tcna]).value = TN
      else:
        TFPN_sheet.cell(row, colNames[tcna]).value = NA

def getSEtable(diagnosis, cnas, TFPN_sheet):
  # Names to indexes of TFPN sheet
  colNames = {}
  for i in range(TFPN_sheet.max_column):
    col = i + 1
    colName = TFPN_sheet.cell(1, col).value
    colNames[colName] = col
  # Create table
  table = {}
  table['names'] = ['Lesion', 'FN', 'FP', 'TN', 'TP', 'Sensitivity', 'Specificity', 'PPV', 'NPV']
  for cna in cnas: 
    tcna = f't{cna}'
    col = colNames[tcna]
    FNcount = 0; FPcount = 0; TNcount = 0; TPcount = 0
    for i in range(200):
      row = i + 2
      if (TFPN_sheet.cell(row, col).value==FN and TFPN_sheet.cell(row, colNames['Diagnosis']).value==diagnosis):
        FNcount = FNcount + 1
      elif (TFPN_sheet.cell(row, col).value==FP and TFPN_sheet.cell(row, colNames['Diagnosis']).value==diagnosis):
        FPcount = FPcount + 1 
      elif (TFPN_sheet.cell(row, col).value==TN and TFPN_sheet.cell(row, colNames['Diagnosis']).value==diagnosis):
        TNcount = TNcount + 1 
      elif (TFPN_sheet.cell(row, col).value==TP and TFPN_sheet.cell(row, colNames['Diagnosis']).value==diagnosis):
        TPcount = TPcount + 1 
    table[tcna] = [cna, FNcount, FPcount, TNcount, TPcount, TPcount / (TPcount + FNcount), TNcount/(TNcount + FPcount), TPcount/(TPcount + FPcount), TNcount/(TNcount + FNcount)] 
  return table

def writeSETable(diagnosis, cnas, TFPN_sheet, tableFileName):
  tableSE = getSEtable(diagnosis, cnas, TFPN_sheet)
  #print(f'method {method}') 
  print(f'Writing table {tableFileName}')
  with open(tableFileName, 'w') as f:
    for item in tableSE['names']:
      f.write(str(item))
      f.write(' ')
    f.write('\n')
    for cna in cnas:
      tcna = 't'+cna
      for item in tableSE[tcna]:
        f.write(str(item))
        f.write(' ')
      f.write('\n')

ref_table_name = 'tables/CNV_allcases_r.xlsx'
ref_book = openpyxl.load_workbook(ref_table_name, read_only=True)
ref_sheet = ref_book.active

cnvkit_segment_cn_table_name = 'tables/CNV_copynumber_cnvkit_segment.xlsx'
cnvkit_segment_cn_book = openpyxl.load_workbook(cnvkit_segment_cn_table_name)
cnvkit_segment_cn_sheet = cnvkit_segment_cn_book.active

cnvkit_segment_TFPN_table_name = 'tables/CNV_TFPN_cnvkit_segment.xlsx'
cnvkit_segment_TFPN_book = openpyxl.load_workbook(cnvkit_segment_TFPN_table_name)
cnvkit_segment_TFPN_sheet = cnvkit_segment_TFPN_book.active

cnvkit_region_cn_table_name = 'tables/CNV_copynumber_cnvkit_region.xlsx'
cnvkit_region_cn_book = openpyxl.load_workbook(cnvkit_region_cn_table_name)
cnvkit_region_cn_sheet = cnvkit_region_cn_book.active

cnvkit_region_TFPN_table_name = 'tables/CNV_TFPN_cnvkit_region.xlsx'
cnvkit_region_TFPN_book = openpyxl.load_workbook(cnvkit_region_TFPN_table_name)
cnvkit_region_TFPN_sheet = cnvkit_region_TFPN_book.active

ichorcna_segment_cn_table_name = 'tables/CNV_copynumber_ichorcna_segment.xlsx'
ichorcna_segment_cn_book = openpyxl.load_workbook(ichorcna_segment_cn_table_name)
ichorcna_segment_cn_sheet = ichorcna_segment_cn_book.active

ichorcna_segment_TFPN_table_name = 'tables/CNV_TFPN_ichorcna_segment.xlsx'
ichorcna_segment_TFPN_book = openpyxl.load_workbook(ichorcna_segment_TFPN_table_name)
ichorcna_segment_TFPN_sheet = ichorcna_segment_TFPN_book.active

ichorcna_region_cn_table_name = 'tables/CNV_copynumber_ichorcna_region.xlsx'
ichorcna_region_cn_book = openpyxl.load_workbook(ichorcna_region_cn_table_name)
ichorcna_region_cn_sheet = ichorcna_region_cn_book.active

ichorcna_region_TFPN_table_name = 'tables/CNV_TFPN_ichorcna_region.xlsx'
ichorcna_region_TFPN_book = openpyxl.load_workbook(ichorcna_region_TFPN_table_name)
ichorcna_region_TFPN_sheet = ichorcna_region_TFPN_book.active

Nrows = ref_sheet.max_row
Ncolumns = ref_sheet.max_column

colNames = {}
for i in range(Ncolumns):
  col = i + 1
  colName = ref_sheet.cell(1, col).value
  colNames[colName] = col

# TODO add comments
updateTables = False
if (len(sys.argv)>=2):
  if (sys.argv[1]=='yes'):
    updateTables = True
if (updateTables):
  for i in range(200):#range(Nrows-1):
    row = i + 2 
    HLabel = ref_sheet.cell(row, colNames['HSTAMP_Label']).value
    # cnvkit tables
    cnvkitCallSegFile = f'cnvkit/results-cnn-tumor/Sample_{HLabel}-T1_Tumor.samtools.call.cns'
    w_cns = weightedCN(cnvkitCallSegFile, cnvkitFormat)
    updateSheets(row, colNames, w_cns, cnvkit_segment_cn_sheet, cnvkit_segment_TFPN_sheet)
    cnvkitCallRegFile = f'cnvkit/results-cnn-tumor/Sample_{HLabel}-T1_Tumor.samtools.call.cnr'
    w_cns = weightedCN(cnvkitCallRegFile, cnvkitFormat)
    updateSheets(row, colNames, w_cns, cnvkit_region_cn_sheet, cnvkit_region_TFPN_sheet)
    print(f'cnvkit {HLabel} {w_cns}')
    # ichorcna tables
    ichorcnaSegFile = f'ichorcna/results-ichorcna/{HLabel}.seg'
    w_cns = weightedCN(ichorcnaSegFile, ichorcnaSegmentFormat)
    updateSheets(row, colNames, w_cns, ichorcna_segment_cn_sheet, ichorcna_segment_TFPN_sheet)
    ichorcnaRegFile = f'ichorcna/results-ichorcna/{HLabel}.cna.seg'
    w_cns = weightedCN(ichorcnaRegFile, ichorcnaRegionFormat)
    updateSheets(row, colNames, w_cns, ichorcna_region_cn_sheet, ichorcna_region_TFPN_sheet)
    print(f'ichorcna {HLabel} {w_cns}')
  # Save the updated excel table
  cnvkit_segment_cn_book.save(cnvkit_segment_cn_table_name)
  cnvkit_segment_TFPN_book.save(cnvkit_segment_TFPN_table_name)
  cnvkit_region_cn_book.save(cnvkit_region_cn_table_name)
  cnvkit_region_TFPN_book.save(cnvkit_region_TFPN_table_name)
  ichorcna_segment_cn_book.save(ichorcna_segment_cn_table_name)
  ichorcna_segment_TFPN_book.save(ichorcna_segment_TFPN_table_name)
  ichorcna_region_cn_book.save(ichorcna_region_cn_table_name)
  ichorcna_region_TFPN_book.save(ichorcna_region_TFPN_table_name)

# Writing the resulting tables
TFPN_sheets = {}
TFPN_sheets['cnvkit-segment'] = cnvkit_segment_TFPN_sheet
TFPN_sheets['cnvkit-region'] = cnvkit_region_TFPN_sheet
TFPN_sheets['ichorcna-segment'] = ichorcna_segment_TFPN_sheet
TFPN_sheets['ichorcna-region'] = ichorcna_region_TFPN_sheet
for method in TFPN_sheets:
  # Define CLL diagnosis, the list of its CNAs, and the name of the written table
  diagnosis = 'CLL'; cnas = ['11q', '13q', '17p', 'trisomy12'] 
  tableFileName = 'tables/'+method+'-'+diagnosis+'-tableSE.txt'
  writeSETable(diagnosis, cnas, TFPN_sheets[method], tableFileName)
  # Define MDS diagnosis, the list of its CNAs, and the name of the written table
  diagnosis = 'MDS'; cnas = ['5q', '7q', '20q', 'trisomy8'] 
  tableFileName = 'tables/'+method+'-'+diagnosis+'-tableSE.txt'
  writeSETable(diagnosis, cnas, TFPN_sheets[method], tableFileName)
  # Define MM diagnosis, the list of its CNAs, and the name of the written table
  diagnosis = 'MM'; cnas = ['1p', '1q', '13q', '17p'] 
  tableFileName = 'tables/'+method+'-'+diagnosis+'-tableSE.txt'
  writeSETable(diagnosis, cnas, TFPN_sheets[method], tableFileName)
