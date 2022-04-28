#!/usr/bin/env python3

import openpyxl
import sys
from CNAdefs import *
from weightedmeanvalue import weightedMeanValues

# Clinical result of karyotype or fish
TP = 'TP'; FP = 'FP'; FN = 'FN'; TN = 'TN'; NA = 'NA'
def updateSheets(row, colNames, w_values, value_sheet, TFPN_sheet):
  gainThreshold = 1.96; deletionThreshold = -1.96
  # w_values is a dictionary useing cna keys
  for cna in w_values: 
    value = w_values[cna]
    kcna = f'k{cna}'
    fcna = f'f{cna}'
    tcna = f't{cna}'
    value_sheet.cell(row, colNames[kcna]).value = value
    value_sheet.cell(row, colNames[fcna]).value = value
    value_sheet.cell(row, colNames[tcna]).value = value
    tcna_value = ref_sheet.cell(row, colNames[tcna]).value
    # tcna of loss is t5q, t7q, etc, and gain is t1q, ttrisomy8 , and ttrisomy12 (skip leading 't')
    if (tcna[1:]=='1q' or (tcna[1:]=='trisomy8' or tcna[1:]=='trisomy12')):
      # Gain true-false-positive-negative logic
      if (tcna_value==1 and value > gainThreshold):
        TFPN_sheet.cell(row, colNames[tcna]).value = TP 
      elif (tcna_value==0 and value > gainThreshold):
        TFPN_sheet.cell(row, colNames[tcna]).value = FP
      elif (tcna_value==1 and value <= gainThreshold):
        TFPN_sheet.cell(row, colNames[tcna]).value = FN
      elif (tcna_value==0 and value <= gainThreshold):
        TFPN_sheet.cell(row, colNames[tcna]).value = TN
      else:
        TFPN_sheet.cell(row, colNames[tcna]).value = NA
    else:
      # Deletion true-false-positive-negative logic
      if (tcna_value==1 and value < deletionThreshold):
        TFPN_sheet.cell(row, colNames[tcna]).value = TP 
      elif (tcna_value==0 and value < deletionThreshold):
        TFPN_sheet.cell(row, colNames[tcna]).value = FP
      elif (tcna_value==1 and value >= deletionThreshold):
        TFPN_sheet.cell(row, colNames[tcna]).value = FN
      elif (tcna_value==0 and value >= deletionThreshold):
        TFPN_sheet.cell(row, colNames[tcna]).value = TN
      else:
        TFPN_sheet.cell(row, colNames[tcna]).value = NA

def getSEtable(diagnosis, cnas, TFPN_sheet):
  # Names to indexes of TFPN sheet
  colNames = {}
  for i in range(TFPN_sheet.max_column):
    col = i + 1
    colName = TFPN_sheet.cell(1, col).value
    colNames[colName] = col
  # Create table
  table = {}
  table['names'] = ['Lesion', 'FN', 'FP', 'TN', 'TP', 'Sensitivity', 'Specificity', 'PPV', 'NPV']
  for cna in cnas: 
    tcna = f't{cna}'
    col = colNames[tcna]
    FNcount = 0; FPcount = 0; TNcount = 0; TPcount = 0
    for i in range(200):
      row = i + 2
      if (TFPN_sheet.cell(row, col).value==FN and TFPN_sheet.cell(row, colNames['Diagnosis']).value==diagnosis):
        FNcount = FNcount + 1
      elif (TFPN_sheet.cell(row, col).value==FP and TFPN_sheet.cell(row, colNames['Diagnosis']).value==diagnosis):
        FPcount = FPcount + 1 
      elif (TFPN_sheet.cell(row, col).value==TN and TFPN_sheet.cell(row, colNames['Diagnosis']).value==diagnosis):
        TNcount = TNcount + 1 
      elif (TFPN_sheet.cell(row, col).value==TP and TFPN_sheet.cell(row, colNames['Diagnosis']).value==diagnosis):
        TPcount = TPcount + 1 
    table[tcna] = [cna, FNcount, FPcount, TNcount, TPcount, TPcount / (TPcount + FNcount), TNcount/(TNcount + FPcount), TPcount/(TPcount + FPcount), TNcount/(TNcount + FNcount)] 
  return table

def writeSETable(diagnosis, cnas, TFPN_sheet, tableFileName):
  tableSE = getSEtable(diagnosis, cnas, TFPN_sheet)
  #print(f'method {method}') 
  print(f'Writing table {tableFileName}')
  with open(tableFileName, 'w') as f:
    for item in tableSE['names']:
      f.write(str(item))
      f.write(' ')
    f.write('\n')
    for cna in cnas:
      tcna = 't'+cna
      for item in tableSE[tcna]:
        f.write(str(item))
        f.write(' ')
      f.write('\n')

ref_table_name = 'tables/CNV_allcases_r.xlsx'
ref_book = openpyxl.load_workbook(ref_table_name, read_only=True)
ref_sheet = ref_book.active

canary_z_table_name = 'tables/CNV_copynumber_canary.xlsx'
canary_z_book = openpyxl.load_workbook(canary_z_table_name)
canary_z_sheet = canary_z_book.active

canary_TFPN_table_name = 'tables/CNV_TFPN_canary.xlsx'
canary_TFPN_book = openpyxl.load_workbook(canary_TFPN_table_name)
canary_TFPN_sheet = canary_TFPN_book.active

Nrows = ref_sheet.max_row
Ncolumns = ref_sheet.max_column

colNames = {}
for i in range(Ncolumns):
  col = i + 1
  colName = ref_sheet.cell(1, col).value
  colNames[colName] = col

# TODO add comments
updateTables = False
if (len(sys.argv)>=2):
  if (sys.argv[1]=='yes'):
    updateTables = True
if (updateTables):
  for i in range(200):#range(Nrows-1):
    row = i + 2 
    HLabel = ref_sheet.cell(row, colNames['HSTAMP_Label']).value
    #fileName = 'canary-python/results-canary/results-canary-mse/Sample_HSTAMP0054-T1_Tumor.cnvZscores'
    fileName = 'canary-python/results-canary/results-canary-mse/Sample_{HLabel}-T1_Tumor.cnvZscores'
    # Colun names used in canary output
    chrColName = "#chr"; startColName = "start"; endColName = "end"; valueColName = "gc.corrected.norm.log.std.index.zWeighted.Final"
    w_zscores = weightedMeanValues(fileName, chrColName, startColName, endColName, valueColName)
    updateSheets(row, colNames, w_zscores, canary_z_sheet, canary_TFPN_sheet)
    print(f'canary {HLabel} {w_zscores}')
  # Save the updated excel table
  canary_z_book.save(canary_z_table_name)
  canary_TFPN_book.save(canary_TFPN_table_name)

# Writing the resulting tables
TFPN_sheets = {}
TFPN_sheets['canary'] = canary_TFPN_sheet
for method in TFPN_sheets:
  # Define CLL diagnosis, the list of its CNAs, and the name of the written table
  diagnosis = 'CLL'; cnas = ['11q', '13q', '17p', 'trisomy12'] 
  tableFileName = 'tables/'+method+'-'+diagnosis+'-tableSE.txt'
  writeSETable(diagnosis, cnas, TFPN_sheets[method], tableFileName)
  # Define MDS diagnosis, the list of its CNAs, and the name of the written table
  diagnosis = 'MDS'; cnas = ['5q', '7q', '20q', 'trisomy8'] 
  tableFileName = 'tables/'+method+'-'+diagnosis+'-tableSE.txt'
  writeSETable(diagnosis, cnas, TFPN_sheets[method], tableFileName)
  # Define MM diagnosis, the list of its CNAs, and the name of the written table
  diagnosis = 'MM'; cnas = ['1p', '1q', '13q', '17p'] 
  tableFileName = 'tables/'+method+'-'+diagnosis+'-tableSE.txt'
  writeSETable(diagnosis, cnas, TFPN_sheets[method], tableFileName)