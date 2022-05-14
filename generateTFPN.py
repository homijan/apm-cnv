#!/usr/bin/env python3

import openpyxl
import sys
from CNAdefs import *
from weightedmeanvalue import weightedMeanValues

# Clinical result of karyotype or fish
TP = 'TP'; FP = 'FP'; FN = 'FN'; TN = 'TN'; NA = 'NA'
def updateSheets(row, colNames, w_values, value_sheet, TFPN_sheet, gainThreshold, deletionThreshold): 
  # User provides gain/deletion thresholds based on the data
  # w_values is a dictionary useing cna keys
  for cna in w_values: 
    value = w_values[cna]
    kcna = f'k{cna}'
    fcna = f'f{cna}'
    tcna = f't{cna}'
    value_sheet.cell(row, colNames[kcna]).value = value
    value_sheet.cell(row, colNames[fcna]).value = value
    value_sheet.cell(row, colNames[tcna]).value = value
    tcna_value = ref_sheet.cell(row, colNames[tcna]).value
    # tcna of loss is t5q, t7q, etc, and gain is t1q, ttrisomy8 , and ttrisomy12 (skip leading 't')
    if (tcna[1:]=='1q' or (tcna[1:]=='trisomy8' or tcna[1:]=='trisomy12')):
      # Gain true-false-positive-negative logic
      if (tcna_value==1 and value > gainThreshold):
        TFPN_sheet.cell(row, colNames[tcna]).value = TP 
      elif (tcna_value==0 and value > gainThreshold):
        TFPN_sheet.cell(row, colNames[tcna]).value = FP
      elif (tcna_value==1 and value <= gainThreshold):
        TFPN_sheet.cell(row, colNames[tcna]).value = FN
      elif (tcna_value==0 and value <= gainThreshold):
        TFPN_sheet.cell(row, colNames[tcna]).value = TN
      else:
        TFPN_sheet.cell(row, colNames[tcna]).value = NA
    else:
      # Deletion true-false-positive-negative logic
      if (tcna_value==1 and value < deletionThreshold):
        TFPN_sheet.cell(row, colNames[tcna]).value = TP 
      elif (tcna_value==0 and value < deletionThreshold):
        TFPN_sheet.cell(row, colNames[tcna]).value = FP
      elif (tcna_value==1 and value >= deletionThreshold):
        TFPN_sheet.cell(row, colNames[tcna]).value = FN
      elif (tcna_value==0 and value >= deletionThreshold):
        TFPN_sheet.cell(row, colNames[tcna]).value = TN
      else:
        TFPN_sheet.cell(row, colNames[tcna]).value = NA

def getSEtable(diagnosis, cnas, TFPN_sheet):
  # Names to indexes of TFPN sheet
  colNames = {}
  for i in range(TFPN_sheet.max_column):
    col = i + 1
    colName = TFPN_sheet.cell(1, col).value
    colNames[colName] = col
  # Create table
  table = {}
  table['names'] = ['Lesion', 'FN', 'FP', 'TN', 'TP', 'Sensitivity', 'Specificity', 'PPV', 'NPV']
  for cna in cnas: 
    tcna = f't{cna}'
    col = colNames[tcna]
    FNcount = 0; FPcount = 0; TNcount = 0; TPcount = 0
    for i in range(200):
      row = i + 2
      if (TFPN_sheet.cell(row, col).value==FN and TFPN_sheet.cell(row, colNames['Diagnosis']).value==diagnosis):
        FNcount = FNcount + 1
      elif (TFPN_sheet.cell(row, col).value==FP and TFPN_sheet.cell(row, colNames['Diagnosis']).value==diagnosis):
        FPcount = FPcount + 1 
      elif (TFPN_sheet.cell(row, col).value==TN and TFPN_sheet.cell(row, colNames['Diagnosis']).value==diagnosis):
        TNcount = TNcount + 1 
      elif (TFPN_sheet.cell(row, col).value==TP and TFPN_sheet.cell(row, colNames['Diagnosis']).value==diagnosis):
        TPcount = TPcount + 1 
    eps = 1e-32
    table[tcna] = [cna, FNcount, FPcount, TNcount, TPcount, TPcount / (TPcount + FNcount + eps), TNcount / (TNcount + FPcount + eps), TPcount / (TPcount + FPcount + eps), TNcount / (TNcount + FNcount + eps)] 
  return table

def writeSETable(diagnosis, cnas, TFPN_sheet, tableFileName):
  tableSE = getSEtable(diagnosis, cnas, TFPN_sheet)
  #print(f'method {method}') 
  print(f'Writing table {tableFileName}')
  with open(tableFileName, 'w') as f:
    for item in tableSE['names']:
      f.write(str(item))
      f.write(' ')
    f.write('\n')
    for cna in cnas:
      tcna = 't'+cna
      for item in tableSE[tcna]:
        f.write(str(item))
        f.write(' ')
      f.write('\n')

ref_table_name = 'tables/CNV_allcases_r.xlsx'
ref_book = openpyxl.load_workbook(ref_table_name, read_only=True)
ref_sheet = ref_book.active

z_table_name = 'tables/CNV_zscore.xlsx'
z_book = openpyxl.load_workbook(z_table_name)
z_sheet = z_book.active

TFPN_table_name = 'tables/CNV_TFPN.xlsx'
TFPN_book = openpyxl.load_workbook(TFPN_table_name)
TFPN_sheet = TFPN_book.active

Nrows = ref_sheet.max_row
Ncolumns = ref_sheet.max_column

colNames = {}
for i in range(Ncolumns):
  col = i + 1
  colName = ref_sheet.cell(1, col).value
  colNames[colName] = col

# TODO add comments
# Input parametrs
# First provide one of the cnv methods to be used: canary-kurtz-cytobands, canary-kurtz-arms, canary-mse-cytobands, canary-mse-arms, wisecondor
cnvMethod = 'canary-kurtz-arms'
if (len(sys.argv)>=2):
  cnvMethod = sys.argv[1]
updateTables = False
if (len(sys.argv)>=3):
  if (sys.argv[2]=='yes'):
    updateTables = True
if (updateTables):
  for i in range(200):#range(Nrows-1):
    row = i + 2 
    HLabel = ref_sheet.cell(row, colNames['HSTAMP_Label']).value
    # Default weighted meanvalue (redefine if needed in particular case)
    defaultValue = 'NA'
    # File name and column names used in canary-kurtz output cytobands
    if (cnvMethod=='canary-kurtz-cytobands'):
      fileName = f'/drive3/dkurtz/HEMESTAMP/CANARy/samples/output/Sample_{HLabel}-T1_Tumor.SegmentedGenome.cytobands-noXY.on-off-combined.txt'
      chrColName = "chrNum"; intChromValue = True; startColName = "Start"; endColName = "End"; valueColName = "combinedStoufferZL2CNR"
      gainThreshold = 1.96; deletionThreshold = -1.96
    # File name and column names used in canary-kurtz output arms
    elif (cnvMethod=='canary-kurtz-arms'):
      fileName = f'/drive3/dkurtz/HEMESTAMP/CANARy/samples/output/Sample_{HLabel}-T1_Tumor.SegmentedGenome.arms.on-off-combined.txt'
      chrColName = "chrNum"; intChromValue = True; startColName = "Start"; endColName = "End"; valueColName = "combinedStoufferZL2CNR" 
      gainThreshold = 1.96; deletionThreshold = -1.96
    # File name and column names used in canary-mse output cytobands
    elif (cnvMethod=='canary-mse-cytobands'):
      fileName = f'canary-python/results-canary/results-canary-mse/Sample_{HLabel}-T1_Tumor.cnvZscores'
      chrColName = "#chr"; intChromValue = False; startColName = "start"; endColName = "end"; valueColName = "gc.corrected.norm.log.std.index.zWeighted.Final"
      gainThreshold = 1.96; deletionThreshold = -1.96
    # File name and column names used in canary-mse output arms
    elif (cnvMethod=='canary-mse-arms'):
      fileName = f'/drive3/mse/CNV/Alicia/results-canary3/Sample_{HLabel}-T1_Tumor.cnvZscores'
      chrColName = "#chr"; intChromValue = False; startColName = "start"; endColName = "end"; valueColName = "gc.corrected.norm.log.std.index.zWeighted.Final" 
      gainThreshold = 1.96; deletionThreshold = -1.96
    # File name and column names used in wisecondor output
    elif (cnvMethod=='wisecondor'):
      fileName = f'wisecondor/testSamples/Sample_{HLabel}-T1_Tumor.sorted.samtools-deduped.sorted.offtarget.std.txt'
      chrColName = "chrNum"; intChromValue = True; startColName = "Start"; endColName = "End"; valueColName = "z-score"; defaultValue = 0
      gainThreshold = 1.96; deletionThreshold = -1.96
    # File name and column names used in cnvki-cnst output. Note that cnvkit uses copynumber rather than z-score
    elif (cnvMethod=='cnvkit-cns'):
      fileName = f'cnvkit/results-cnn-tumor/Sample_{HLabel}-T1_Tumor.samtools.call.cns'
      chrColName = "chromosome"; intChromValue = False; startColName = "start"; endColName = "end"; valueColName = "cn"
      gainThreshold = 2.0; deletionThreshold = 2.0
    # File name and column names used in cnvkit-cnr output. Note that cnvkit uses copynumber rather than z-score
    elif (cnvMethod=='cnvkit-cnr'):
      fileName = f'cnvkit/results-cnn-tumor/Sample_{HLabel}-T1_Tumor.samtools.call.cnr'
      chrColName = "chromosome"; intChromValue = False; startColName = "start"; endColName = "end"; valueColName = "cn"
      gainThreshold = 2.0; deletionThreshold = 2.0
   # File name and column names used in ichorcna-cns output. Note that ichorcna uses copynumber rather than z-score
    elif (cnvMethod=='ichorcna-cns'):
      fileName = f'ichorcna/results-ichorcna/{HLabel}.seg'
      chrColName = "chr"; intChromValue = True; startColName = "start"; endColName = "end"; valueColName = "copy.number"
      gainThreshold = 2.0; deletionThreshold = 2.0
   # File name and column names used in ichorcna-cnr output. Note that ichorcna uses copynumber rather than z-score
    elif (cnvMethod=='ichorcna-cnr'):
      fileName = f'ichorcna/results-ichorcna/{HLabel}.cna.seg'
      chrColName = "chr"; intChromValue = True; startColName = "start"; endColName = "end"; valueColName = f'{HLabel}.copy.number'
      gainThreshold = 2.0; deletionThreshold = 2.0
    else:
      print(f'Provided cnv method {cnvMethod} not in the supported list: canary-kurtz-cytobands, canary-kurtz-arms, canary-mse-cytobands, canary-mse-arms, wisecondor, cnvkit-cns, cnvkit-cnr, ichorcna-cns,ichorcna-cnr')

    # Obtain zscore for every cna from CNAdefs
    print(f'fileName: {fileName}')
    w_zscores = weightedMeanValues(fileName, chrColName, startColName, endColName, valueColName, intChromValue, defaultValue)
    updateSheets(row, colNames, w_zscores, z_sheet, TFPN_sheet, gainThreshold, deletionThreshold)
    print(f'{cnvMethod} {HLabel} {w_zscores}')
  # Save the updated excel table
  z_book.save(z_table_name)
  TFPN_book.save(TFPN_table_name)

# Writing the resulting tables
TFPN_sheets = {}
TFPN_sheets[cnvMethod] = TFPN_sheet
for method in TFPN_sheets:
  # Define CLL diagnosis, the list of its CNAs, and the name of the written table
  diagnosis = 'CLL'; cnas = ['11q', '13q', '17p', 'trisomy12'] 
  tableFileName = 'tables/'+method+'-'+diagnosis+'-tableSE.txt'
  writeSETable(diagnosis, cnas, TFPN_sheets[method], tableFileName)
  # Define MDS diagnosis, the list of its CNAs, and the name of the written table
  diagnosis = 'MDS'; cnas = ['5q', '7q', '20q', 'trisomy8'] 
  tableFileName = 'tables/'+method+'-'+diagnosis+'-tableSE.txt'
  writeSETable(diagnosis, cnas, TFPN_sheets[method], tableFileName)
  # Define MM diagnosis, the list of its CNAs, and the name of the written table
  diagnosis = 'MM'; cnas = ['1p', '1q', '13q', '17p'] 
  tableFileName = 'tables/'+method+'-'+diagnosis+'-tableSE.txt'
  writeSETable(diagnosis, cnas, TFPN_sheets[method], tableFileName)
